"""Plantilla Excel de parámetros: descarga (prefijada con los valores actuales)
e importación masiva. Hojas: Regulados, Generales, SSAA."""
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .models import ParametroGeneral, ParametroRegulado, ProfileFactor, SerieSSAA, Tarifa

lado = Side(style="thin", color="B0B7C3")
BORDE = Border(left=lado, right=lado, top=lado, bottom=lado)
F_CAB = Font(name="Arial", size=9, bold=True, color="FFFFFF")
FILL_CAB = PatternFill("solid", start_color="4D4D4F")
F_INPUT = Font(name="Arial", size=10, color="0000C0")

MESES_NOMBRE = {i: n for i, n in enumerate(
    ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], start=1)}
MESES_NUMERO = {n.lower(): i for i, n in MESES_NOMBRE.items()}


def _cabecera(ws, textos, anchos):
    for j, (texto, ancho) in enumerate(zip(textos, anchos), start=1):
        c = ws.cell(row=1, column=j, value=texto)
        c.font = F_CAB
        c.fill = FILL_CAB
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = ancho
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _fila(ws, f, valores, formatos=None):
    for j, v in enumerate(valores, start=1):
        c = ws.cell(row=f, column=j, value=v)
        c.font = F_INPUT
        c.border = BORDE
        if formatos and formatos.get(j):
            c.number_format = formatos[j]


def generar_plantilla():
    """Plantilla con los parámetros actuales como punto de partida."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Regulados"
    _cabecera(ws, ["Tipo", "Término", "Tarifa", "Vigencia inicio", "Vigencia fin (vacío = vigente)",
                   "P1", "P2", "P3", "P4", "P5", "P6", "Fuente"],
              [10, 10, 8, 14, 14, 12, 12, 12, 12, 12, 12, 45])
    formatos = {4: "DD/MM/YYYY", 5: "DD/MM/YYYY", **{j: "0.000000" for j in range(6, 12)}}
    f = 2
    for p in ParametroRegulado.objects.order_by("tarifa", "tipo", "termino", "-vigencia_inicio"):
        _fila(ws, f, [p.tipo, p.termino, p.tarifa, p.vigencia_inicio, p.vigencia_fin,
                      p.p1, p.p2, p.p3, p.p4, p.p5, p.p6, p.fuente], formatos)
        f += 1
    dv_tipo = DataValidation(type="list", formula1='"peaje,cargo,perdidas,imp_local"', allow_blank=True)
    dv_term = DataValidation(type="list", formula1='"energia,potencia"', allow_blank=True)
    dv_tarifa = DataValidation(type="list", formula1='"2.0TD,3.0TD,6.1TD,6.2TD,6.3TD,6.4TD"', allow_blank=True)
    for dv, col in ((dv_tipo, "A"), (dv_term, "B"), (dv_tarifa, "C")):
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    ws = wb.create_sheet("Generales")
    _cabecera(ws, ["Clave", "Valor (tanto por uno)", "Vigencia inicio", "Vigencia fin", "Fuente"],
              [10, 18, 14, 14, 45])
    formatos = {2: "0.000000000", 3: "DD/MM/YYYY", 4: "DD/MM/YYYY"}
    f = 2
    for p in ParametroGeneral.objects.order_by("clave", "-vigencia_inicio"):
        _fila(ws, f, [p.clave, p.valor, p.vigencia_inicio, p.vigencia_fin, p.fuente], formatos)
        f += 1
    dv_clave = DataValidation(type="list", formula1='"iee"', allow_blank=True)
    ws.add_data_validation(dv_clave)
    dv_clave.add("A2:A100")

    ws = wb.create_sheet("SSAA")
    _cabecera(ws, ["Mes (1-12 o nombre)", "SSAA estimado (€/MWh)", "SSAA real (€/MWh, opcional)", "Año (informativo)"],
              [18, 22, 24, 16])
    formatos = {2: "0.000", 3: "0.000", 4: "0"}
    f = 2
    for s in SerieSSAA.objects.order_by("mes"):
        _fila(ws, f, [MESES_NOMBRE[s.mes], s.valor_considerado, s.valor_real, s.anio], formatos)
        f += 1

    ws = wb.create_sheet("ProfileFactors")
    _cabecera(ws, ["Ámbito", "Mes (1-12 o nombre)", "P1", "P2", "P3", "P4", "P5", "P6", "Año (informativo)"],
              [16, 18, 10, 10, 10, 10, 10, 10, 16])
    formatos = {**{c: "0.0000" for c in range(3, 9)}, 9: "0"}
    f = 2
    for pf in ProfileFactor.objects.order_by("ambito", "mes"):
        _fila(ws, f, [pf.ambito, MESES_NOMBRE[pf.mes], pf.p1, pf.p2, pf.p3, pf.p4, pf.p5, pf.p6, pf.anio], formatos)
        f += 1
    dv_amb = DataValidation(type="list", formula1='"peninsula_3_6,peninsula_20,canarias,baleares"', allow_blank=True)
    ws.add_data_validation(dv_amb)
    dv_amb.add("A2:A200")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fecha(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    raise ValueError(f"«{valor}» no es una fecha")


def _num(valor, defecto=None):
    if valor is None or valor == "":
        return defecto
    return Decimal(str(valor))


def importar_parametros(fichero):
    """Importa la plantilla. Devuelve (creados, actualizados, errores).
    Si hay errores no se guarda nada (la vista lo envuelve en transacción)."""
    errores = []
    creados = actualizados = 0
    try:
        wb = load_workbook(fichero, data_only=True, read_only=True)
    except Exception:
        return 0, 0, ["El archivo no es un Excel válido (.xlsx)."]

    if "Regulados" in wb.sheetnames:
        for n, fila in enumerate(wb["Regulados"].iter_rows(min_row=2, max_col=12, values_only=True), start=2):
            if not any(v is not None and v != "" for v in fila):
                continue
            tipo, termino, tarifa = (str(v).strip().lower() if v else "" for v in fila[:3])
            ctx = f"Regulados fila {n}"
            if tipo not in ("peaje", "cargo", "perdidas", "imp_local"):
                errores.append(f"{ctx}: tipo «{fila[0]}» no válido (peaje/cargo/perdidas/imp_local).")
                continue
            if termino not in ("energia", "potencia"):
                errores.append(f"{ctx}: término «{fila[1]}» no válido (energia/potencia).")
                continue
            tarifa = str(fila[2]).strip()
            if tarifa not in Tarifa.values:
                errores.append(f"{ctx}: tarifa «{fila[2]}» no válida.")
                continue
            try:
                inicio = _fecha(fila[3])
                fin = _fecha(fila[4])
                valores = [_num(v, Decimal(0)) for v in fila[5:11]]
            except (ValueError, InvalidOperation) as e:
                errores.append(f"{ctx}: {e}.")
                continue
            if inicio is None:
                errores.append(f"{ctx}: falta la vigencia de inicio.")
                continue
            _, nuevo = ParametroRegulado.objects.update_or_create(
                tipo=tipo, termino=termino, tarifa=tarifa, vigencia_inicio=inicio,
                defaults={
                    "vigencia_fin": fin,
                    "p1": valores[0], "p2": valores[1], "p3": valores[2],
                    "p4": valores[3], "p5": valores[4], "p6": valores[5],
                    "fuente": str(fila[11] or ""), "validado": False,
                },
            )
            creados += nuevo
            actualizados += not nuevo

    if "Generales" in wb.sheetnames:
        for n, fila in enumerate(wb["Generales"].iter_rows(min_row=2, max_col=5, values_only=True), start=2):
            if not any(v is not None and v != "" for v in fila):
                continue
            ctx = f"Generales fila {n}"
            clave = str(fila[0] or "").strip().lower()
            if clave not in ParametroGeneral.Clave.values:
                errores.append(f"{ctx}: clave «{fila[0]}» no válida.")
                continue
            try:
                valor = _num(fila[1])
                inicio = _fecha(fila[2])
                fin = _fecha(fila[3])
            except (ValueError, InvalidOperation) as e:
                errores.append(f"{ctx}: {e}.")
                continue
            if valor is None or inicio is None:
                errores.append(f"{ctx}: faltan el valor o la vigencia de inicio.")
                continue
            _, nuevo = ParametroGeneral.objects.update_or_create(
                clave=clave, vigencia_inicio=inicio,
                defaults={"valor": valor, "vigencia_fin": fin,
                          "fuente": str(fila[4] or ""), "validado": False},
            )
            creados += nuevo
            actualizados += not nuevo

    def _mes(bruto):
        if isinstance(bruto, (int, float)) and 1 <= int(bruto) <= 12:
            return int(bruto)
        return MESES_NUMERO.get(str(bruto or "").strip().lower())

    def _anio(v):
        try:
            return int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    if "SSAA" in wb.sheetnames:
        for n, fila in enumerate(wb["SSAA"].iter_rows(min_row=2, max_col=4, values_only=True), start=2):
            if not any(v is not None and v != "" for v in fila):
                continue
            ctx = f"SSAA fila {n}"
            mes = _mes(fila[0])
            if mes is None:
                errores.append(f"{ctx}: mes «{fila[0]}» no válido.")
                continue
            try:
                considerado = _num(fila[1])
                real = _num(fila[2])
            except InvalidOperation:
                errores.append(f"{ctx}: valores SSAA no numéricos.")
                continue
            if considerado is None:
                errores.append(f"{ctx}: falta el SSAA estimado.")
                continue
            _, nuevo = SerieSSAA.objects.update_or_create(
                mes=mes,
                defaults={"valor_considerado": considerado, "valor_real": real, "anio": _anio(fila[3])},
            )
            creados += nuevo
            actualizados += not nuevo

    if "ProfileFactors" in wb.sheetnames:
        ambitos_validos = set(ProfileFactor.Ambito.values)
        for n, fila in enumerate(wb["ProfileFactors"].iter_rows(min_row=2, max_col=9, values_only=True), start=2):
            if not any(v is not None and v != "" for v in fila):
                continue
            ctx = f"ProfileFactors fila {n}"
            ambito = str(fila[0] or "").strip()
            if ambito not in ambitos_validos:
                errores.append(f"{ctx}: ámbito «{fila[0]}» no válido.")
                continue
            mes = _mes(fila[1])
            if mes is None:
                errores.append(f"{ctx}: mes «{fila[1]}» no válido.")
                continue
            try:
                valores = [_num(fila[i]) for i in range(2, 8)]
            except InvalidOperation:
                errores.append(f"{ctx}: profile factors no numéricos.")
                continue
            _, nuevo = ProfileFactor.objects.update_or_create(
                ambito=ambito, mes=mes,
                defaults={
                    "p1": valores[0], "p2": valores[1], "p3": valores[2],
                    "p4": valores[3], "p5": valores[4], "p6": valores[5],
                    "anio": _anio(fila[8]),
                },
            )
            creados += nuevo
            actualizados += not nuevo

    return creados, actualizados, errores
