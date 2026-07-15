"""Importador de la Plantilla_Estudio_Renovacion_v2.xlsx.

Lee las hojas Cliente / Suministros / Consumos, valida y devuelve los datos
listos para crear el expediente. No toca la base de datos: eso lo hace la vista
dentro de una transacción.
"""
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Tarifa

MESES = {n.lower(): i for i, n in enumerate(
    ["enero", "febrero", "marzo", "abril", "mayo", "junio",
     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"], start=1)}

LETRAS_CONTROL_CUPS = "TRWAGMYFPDXBNJZSQVHLCKE"

N_CUPS = 50
FILA_INI_SUMINISTROS = 5
BASE_CONSUMOS = 4
PASO_CONSUMOS = 18


@dataclass
class DatosCliente:
    razon_social: str = ""
    cif: str = ""
    direccion: str = ""
    gestor: str = ""
    gestor_email: str = ""


@dataclass
class DatosPunto:
    fila: int = 0
    titular: str = ""
    cif_titular: str = ""
    cups: str = ""
    direccion: str = ""
    tarifa: str = ""
    potencias: list = field(default_factory=lambda: [None] * 6)
    comercializadora: str = ""
    modalidad: str = "fijo"
    fecha_fin: date | None = None
    energia_peajes_incluidos: bool = True
    precios_energia: list = field(default_factory=lambda: [None] * 6)
    potencia_segun_atr: bool = True
    precios_potencia: list = field(default_factory=lambda: [None] * 6)
    consumos: list = field(default_factory=list)  # [(anio, mes, [p1..p6])]


@dataclass
class ResultadoImportacion:
    cliente: DatosCliente = field(default_factory=DatosCliente)
    puntos: list = field(default_factory=list)
    errores: list = field(default_factory=list)
    avisos: list = field(default_factory=list)

    @property
    def valido(self):
        return not self.errores and self.puntos


def _texto(valor):
    return str(valor).strip() if valor is not None else ""


def _decimal(valor, contexto, errores):
    if valor is None or valor == "":
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        errores.append(f"{contexto}: el valor «{valor}» no es un número válido.")
        return None


def _si_no(valor, defecto):
    texto = _texto(valor).lower()
    if texto in ("sí", "si"):
        return True
    if texto == "no":
        return False
    return defecto


def validar_cups(cups):
    """Valida formato y letras de control del CUPS español (ES + 16 dígitos + 2 letras)."""
    cups = cups.upper().replace(" ", "")
    if len(cups) not in (20, 22) or not cups.startswith("ES"):
        return False
    digitos, letras = cups[2:18], cups[18:20]
    if not digitos.isdigit():
        return False
    resto = int(digitos) % 529
    cociente, modulo = divmod(resto, 23)
    return letras == LETRAS_CONTROL_CUPS[cociente] + LETRAS_CONTROL_CUPS[modulo]


def importar_plantilla(fichero, requiere_cliente=True):
    """Parsea el fichero subido y devuelve un ResultadoImportacion.

    requiere_cliente=False cuando se importa sobre un expediente ya existente
    (no hace falta rellenar la hoja Cliente; solo se usan Suministros y Consumos).
    """
    r = ResultadoImportacion()
    try:
        wb = load_workbook(fichero, data_only=True, read_only=True)
    except Exception:
        r.errores.append("El archivo no es un Excel válido (.xlsx).")
        return r

    for hoja in ("Cliente", "Suministros", "Consumos"):
        if hoja not in wb.sheetnames:
            r.errores.append(f"Falta la hoja «{hoja}». ¿Es la plantilla v2?")
    if r.errores:
        return r

    # ---- Hoja Cliente (etiquetas en B4:B8, valores en C4:C8) ----
    ws = wb["Cliente"]
    valores = [ws.cell(row=f, column=3).value for f in range(4, 9)]
    r.cliente = DatosCliente(
        razon_social=_texto(valores[0]),
        cif=_texto(valores[1]),
        direccion=_texto(valores[2]),
        gestor=_texto(valores[3]),
        gestor_email=_texto(valores[4]),
    )
    if requiere_cliente:
        if not r.cliente.razon_social:
            r.errores.append("Hoja Cliente: falta la razón social.")
        if not r.cliente.cif:
            r.errores.append("Hoja Cliente: falta el CIF/NIF.")

    # ---- Hoja Suministros ----
    ws = wb["Suministros"]
    filas = list(ws.iter_rows(min_row=FILA_INI_SUMINISTROS,
                              max_row=FILA_INI_SUMINISTROS + N_CUPS - 1,
                              min_col=1, max_col=29, values_only=True))
    indice_por_orden = {}
    for n_fila, fila in enumerate(filas, start=FILA_INI_SUMINISTROS):
        cups = _texto(fila[3])
        if not cups:
            continue
        p = DatosPunto(fila=n_fila)
        ctx = f"Suministros fila {n_fila} ({cups})"
        p.cups = cups.upper().replace(" ", "")
        if not validar_cups(p.cups):
            r.errores.append(f"{ctx}: el CUPS no es válido (formato o letras de control).")
        p.titular = _texto(fila[1])
        p.cif_titular = _texto(fila[2])
        p.direccion = _texto(fila[4])
        p.tarifa = _texto(fila[5])
        if p.tarifa not in Tarifa.values:
            r.errores.append(f"{ctx}: tarifa «{p.tarifa}» no reconocida.")
        p.potencias = [_decimal(fila[6 + i], f"{ctx} potencia P{i+1}", r.errores) for i in range(6)]
        p.comercializadora = _texto(fila[12])
        modalidad = _texto(fila[13]).lower()
        p.modalidad = "indexado" if "index" in modalidad else "fijo"
        if p.modalidad == "indexado":
            r.avisos.append(f"{ctx}: contrato actual indexado — en esta fase no se calculará su línea base.")
        fecha = fila[14]
        if isinstance(fecha, datetime):
            p.fecha_fin = fecha.date()
        elif isinstance(fecha, date):
            p.fecha_fin = fecha
        p.energia_peajes_incluidos = _si_no(fila[15], True)
        p.precios_energia = [_decimal(fila[16 + i], f"{ctx} precio energía P{i+1}", r.errores) for i in range(6)]
        p.potencia_segun_atr = _si_no(fila[22], True)
        p.precios_potencia = [_decimal(fila[23 + i], f"{ctx} precio potencia P{i+1}", r.errores) for i in range(6)]

        # Coherencia tarifa ↔ periodos y datos mínimos
        n_periodos_energia = 3 if p.tarifa == "2.0TD" else 6
        n_periodos_potencia = 2 if p.tarifa == "2.0TD" else 6
        if not any(p.potencias[:n_periodos_potencia]):
            r.errores.append(f"{ctx}: faltan las potencias contratadas.")
        if not any(v is not None for v in p.precios_energia[:n_periodos_energia]):
            r.errores.append(f"{ctx}: faltan los precios de energía del contrato actual.")
        if not p.potencia_segun_atr and not any(v is not None for v in p.precios_potencia[:n_periodos_potencia]):
            r.errores.append(
                f"{ctx}: indica que la potencia NO va según ATR pero no hay precios de potencia."
            )
        if p.potencia_segun_atr and any(v is not None for v in p.precios_potencia):
            r.avisos.append(f"{ctx}: potencia según ATR — se ignoran los precios de potencia escritos.")
            p.precios_potencia = [None] * 6
        if p.tarifa == "2.0TD":
            if any(v for v in p.potencias[2:]):
                r.avisos.append(f"{ctx}: 2.0TD solo usa 2 periodos de potencia — se ignoran P3–P6.")
                p.potencias = p.potencias[:2] + [None] * 4
            if any(v is not None for v in p.precios_energia[3:]):
                r.avisos.append(f"{ctx}: 2.0TD solo usa 3 periodos de energía — se ignoran P4–P6.")
                p.precios_energia = p.precios_energia[:3] + [None] * 3
        elif p.tarifa in ("3.0TD", "6.1TD", "6.2TD", "6.3TD", "6.4TD"):
            potencias = [v for v in p.potencias if v is not None]
            if len(potencias) == 6 and any(potencias[i] > potencias[i + 1] for i in range(5)):
                r.avisos.append(f"{ctx}: las potencias contratadas no son crecientes P1≤…≤P6 — revísalo.")

        indice_por_orden[len(indice_por_orden)] = p
        r.puntos.append(p)

    if not r.puntos:
        r.errores.append("La hoja Suministros no contiene ningún CUPS.")
        return r

    cups_vistos = {}
    for p in r.puntos:
        if p.cups in cups_vistos:
            r.errores.append(f"CUPS duplicado en la plantilla: {p.cups}.")
        cups_vistos[p.cups] = True

    # ---- Hoja Consumos (bloque i ↔ suministro i por orden) ----
    ws = wb["Consumos"]
    celdas = {(c.row, c.column): c.value for fila_ in ws.iter_rows() for c in fila_ if c.value is not None}
    for orden, p in indice_por_orden.items():
        r0 = BASE_CONSUMOS + orden * PASO_CONSUMOS
        ctx = f"Consumos de {p.cups}"
        for m in range(12):
            f = r0 + 2 + m
            mes_txt = _texto(celdas.get((f, 1))).lower()
            anio = celdas.get((f, 2))
            kwh = [celdas.get((f, col)) for col in range(3, 9)]
            if not any(v is not None for v in kwh):
                continue
            if mes_txt not in MESES:
                r.errores.append(f"{ctx}: mes «{mes_txt}» no reconocido en la fila {f}.")
                continue
            if not isinstance(anio, (int, float)):
                r.errores.append(f"{ctx}: falta el año del mes de {mes_txt}.")
                continue
            valores = [_decimal(v, f"{ctx} {mes_txt} P{i+1}", r.errores) or Decimal(0)
                       for i, v in enumerate(kwh)]
            p.consumos.append((int(anio), MESES[mes_txt], valores))
        if len(p.consumos) == 0:
            r.errores.append(f"{ctx}: no hay ningún mes de consumo.")
        elif len(p.consumos) < 12:
            r.avisos.append(f"{ctx}: solo {len(p.consumos)} meses de consumo (lo ideal son 12).")
        meses_vistos = {}
        for anio, mes, _ in p.consumos:
            if (anio, mes) in meses_vistos:
                r.errores.append(f"{ctx}: el mes {mes}/{anio} está repetido.")
            meses_vistos[(anio, mes)] = True

    return r
