"""Exportación del comparativo a Excel (openpyxl) y PDF (xhtml2pdf)."""
import io
import os
import tempfile

from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from .consumo import resumen_consumo
from .graficos import grafico_consumo_apilado

LOGO = "static/img/logo-geype.png"
LOGO_RATIO = 200 / 117   # ancho/alto reales del logo
PX_POR_CARACTER = 7      # ancho de columna de Excel (unidades de carácter) -> píxeles
EMU_POR_PX = 9525        # unidad interna de las imágenes de Excel

GRIS = "4D4D4F"
ROJO = "BE1E2D"
VERDE = "2E7D32"
VERDE_CLARO = "EAF5EA"
GRIS_CLARO = "F0F0F2"
ACTUAL = "F7F7F8"
AZUL = "1F3864"        # bandas de sección de las fichas de consumo
VERDE_OSC = "375623"   # cabecera del cuadro de suministros
AMARILLO = "FFF2CC"    # celdas de dato aportado por el cliente
# Un color por periodo tarifario, en gama pastel (mismos tonos que el gráfico del PDF)
COLOR_PERIODO = ["E4959E", "F2B880", "F0DC96", "A8D5B5", "9EC5DE", "B9A7D4"]

lado = Side(style="thin", color="B0B7C3")
BORDE = Border(left=lado, right=lado, top=lado, bottom=lado)
F_CAB = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_TIT = Font(name="Arial", size=13, bold=True, color="FFFFFF")
F_NEG = Font(name="Arial", size=10, bold=True)
F_NOR = Font(name="Arial", size=10)
F_DATO = Font(name="Arial", size=10, bold=True, color="1F3864")
F_VERDE = Font(name="Arial", size=10, bold=True, color=VERDE)
F_ROJO = Font(name="Arial", size=10, color=ROJO)
F_INC = Font(name="Arial", size=8, italic=True, color="6B7280")
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
IZQ = Alignment(horizontal="left", vertical="center", wrap_text=True)
IZQ_1L = Alignment(horizontal="left", vertical="center")   # sin envolver: celdas combinadas
TEXTO_INCLUIDO = "incluido precio energía"
FMT_EUR = "#,##0.00"
FMT_KWH = "#,##0.00000"
FMT_KWH_ENT = "#,##0"
FMT_KW = "#,##0.00"
FMT_PCT = '#,##0.0" %"'
FMT_PCT0 = "0.0%"


def _eur(x):
    """16894.56 -> '16.894,56' (formato español)."""
    return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _set(ws, fila, col, valor, *, font=F_NOR, fill=None, fmt=None, align=None):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = font
    c.border = BORDE
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    return c


def _tabla_comparativo(ws, fila0, comparativo, por_punto=None):
    """Pinta la tabla comparativa con la mejor oferta resaltada. por_punto: pk o None."""
    def desglose(col):
        return col["por_punto"].get(por_punto) if por_punto else col["agregado"]

    def actual():
        return comparativo["actual_por_punto"].get(por_punto) if por_punto else comparativo["actual"]

    columnas = comparativo["columnas"]

    # Cabecera
    _set(ws, fila0, 1, "", font=F_CAB, fill=GRIS, align=CENTRO)
    _set(ws, fila0, 2, "Contrato actual", font=F_CAB, fill=GRIS, align=CENTRO)
    for j, col in enumerate(columnas, start=3):
        etiqueta = col["oferta"].etiqueta
        if col.get("ranking"):
            etiqueta = f"{col['ranking']}º  {etiqueta}"
        if col["oferta"].gdo:
            etiqueta += " (GdO)"
        _set(ws, fila0, j, etiqueta, font=F_CAB, fill=(VERDE if col.get("mejor") else GRIS), align=CENTRO)

    act = actual()
    # marca: None fila normal; "concepto:<nombre>" muestra "incluido precio energía"
    # cuando la oferta no lleva ese concepto; "ssaa" cuando la oferta lo incluye en el precio.
    filas = [
        ("ATR potencia (€)", lambda d: d.potencia, FMT_EUR, False, None),
        ("Energía (€)", lambda d: d.energia, FMT_EUR, False, None),
    ]
    for nombre in comparativo["conceptos_nombres"]:
        filas.append((f"{nombre} (€)", lambda d, n=nombre: d.conceptos.get(n), FMT_EUR, False, f"concepto:{nombre}"))
    filas += [
        ("Regularización SSAA (€)", lambda d: d.ssaa, FMT_EUR, False, "ssaa"),
        ("IEE (€)", lambda d: d.iee, FMT_EUR, False, None),
        ("TOTAL anual (€)", lambda d: d.total, FMT_EUR, True, None),
        ("€/kWh medio", lambda d: d.eur_kwh, FMT_KWH, False, None),
    ]

    def _celda(d, col, fn, marca):
        """Devuelve el valor a escribir: número, texto 'incluido…' o '—'."""
        if d is None:
            return "—"
        if marca == "ssaa" and col is not None and col["oferta"].ssaa_tipo == "incluido":
            return TEXTO_INCLUIDO
        v = fn(d)
        if marca and marca.startswith("concepto:") and v is None:
            return TEXTO_INCLUIDO
        return float(v) if v is not None else "—"

    f = fila0
    for etiqueta, fn, fmt, es_total, marca in filas:
        f += 1
        _set(ws, f, 1, etiqueta, font=F_NEG, fill=(GRIS_CLARO if es_total else None))
        va = _celda(act, None, fn, marca) if act else "—"
        es_texto = va == TEXTO_INCLUIDO
        _set(ws, f, 2, va, font=(F_INC if es_texto else (F_NEG if es_total else F_NOR)),
             fill=ACTUAL, fmt=(fmt if isinstance(va, float) else None),
             align=CENTRO if es_texto else None)
        for j, col in enumerate(columnas, start=3):
            d = desglose(col)
            v = _celda(d, col, fn, marca)
            es_texto = v == TEXTO_INCLUIDO
            _set(ws, f, j, v, font=(F_INC if es_texto else (F_NEG if es_total else F_NOR)),
                 fill=(VERDE_CLARO if col.get("mejor") else None),
                 fmt=(fmt if isinstance(v, float) else None),
                 align=CENTRO if es_texto else None)

    if not por_punto:
        for etiqueta, clave, fmt in [("Ahorro (€)", "ahorro", FMT_EUR), ("Variación (%)", "ahorro_pct", FMT_PCT)]:
            f += 1
            _set(ws, f, 1, etiqueta, font=F_NEG)
            _set(ws, f, 2, "—", fill=ACTUAL, align=CENTRO)
            for j, col in enumerate(columnas, start=3):
                v = col.get(clave)
                fill = VERDE_CLARO if col.get("mejor") else None
                if v is None:
                    _set(ws, f, j, "—", fill=fill, align=CENTRO)
                else:
                    font = F_VERDE if v < 0 else (F_ROJO if v > 0 else F_NOR)
                    _set(ws, f, j, float(v), font=font, fill=fill, fmt=fmt)
    return f


def _ancho_columnas(ws, n_cols, fila_ref):
    ws.column_dimensions["A"].width = 27
    for j in range(2, n_cols + 1):
        ws.column_dimensions[ws.cell(row=fila_ref, column=j).column_letter].width = 17


def comparativo_a_excel(expediente, comparativo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    n_cols = 2 + len(comparativo["columnas"])
    ancho = max(n_cols, 4)

    # Logo corporativo centrado en la cabecera
    _ancho_columnas(ws, n_cols, 8)
    _logo_centrado(ws, ancho)
    for r in (2, 3):
        ws.row_dimensions[r].height = 12

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ancho)
    _set(ws, 4, 1, f"ESTUDIO DE RENOVACIÓN — {expediente.cliente_razon_social}",
         font=F_TIT, fill=GRIS, align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[4].height = 26

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ancho)
    ws.cell(row=5, column=1,
            value=f"{expediente.codigo} · CIF {expediente.cliente_cif} · "
                  f"a fecha {comparativo['fecha']:%d/%m/%Y} · importes anuales sin IVA").font = F_NOR

    # Banner con la mejor oferta
    mejor = next((c for c in comparativo["columnas"] if c.get("mejor")), None)
    if mejor:
        texto = f"★ Mejor oferta: {mejor['oferta'].etiqueta} — {_eur(mejor['agregado'].total)} €/año"
        if mejor["ahorro"] is not None and mejor["ahorro"] < 0:
            texto += f"  ·  ahorro de {_eur(abs(mejor['ahorro']))} € ({mejor['ahorro_pct']:.1f} %) frente al actual"
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ancho)
        _set(ws, 6, 1, texto, font=Font(name="Arial", size=11, bold=True, color=VERDE), fill=VERDE_CLARO,
             align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[6].height = 20

    f = _tabla_comparativo(ws, 8, comparativo)
    ws.freeze_panes = "B9"

    f += 2
    for col in comparativo["columnas"]:
        if col["oferta"].observaciones:
            ws.cell(row=f, column=1, value=f"{col['oferta'].etiqueta}: {col['oferta'].observaciones}").font = F_NOR
            f += 1
    if comparativo["avisos"]:
        f += 1
        ws.cell(row=f, column=1, value="Avisos del cálculo:").font = F_NEG
        for a in comparativo["avisos"]:
            f += 1
            ws.cell(row=f, column=1, value=f"• {a}").font = F_NOR

    if len(comparativo["puntos"]) > 1:
        for punto in comparativo["puntos"]:
            hoja = wb.create_sheet(punto.cups[-10:])
            hoja.sheet_view.showGridLines = False
            _ancho_columnas(hoja, n_cols, 3)
            hoja.cell(row=1, column=1, value=f"{punto.cups} ({punto.tarifa}) — {punto.direccion}").font = F_NEG
            _tabla_comparativo(hoja, 3, comparativo, por_punto=punto.pk)
            hoja.freeze_panes = "B4"

    # Datos de partida del estudio (suministros y consumos), ya formateados.
    _hojas_datos(wb, expediente)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _logo_centrado(ws, n_cols, alto_px=76):
    """Coloca el logo arriba, centrado sobre las `n_cols` primeras columnas.

    openpyxl solo ancla imágenes a una celda, así que se calcula el desplazamiento
    en EMU desde el borde de la columna para que quede realmente centrado.
    """
    ruta = settings.BASE_DIR / LOGO
    if not ruta.exists():
        return
    ancho_px = round(alto_px * LOGO_RATIO)

    anchos = [(ws.column_dimensions[get_column_letter(j)].width or 8.43) * PX_POR_CARACTER
              for j in range(1, n_cols + 1)]
    izquierda = max(0, (sum(anchos) - ancho_px) / 2)

    col, acumulado = 0, 0.0
    while col < len(anchos) - 1 and acumulado + anchos[col] <= izquierda:
        acumulado += anchos[col]
        col += 1
    desplazamiento = izquierda - acumulado

    img = XLImage(str(ruta))
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col, colOff=int(desplazamiento * EMU_POR_PX), row=0, rowOff=0),
        ext=XDRPositiveSize2D(cx=ancho_px * EMU_POR_PX, cy=alto_px * EMU_POR_PX),
    )
    ws.add_image(img)
    ws.row_dimensions[1].height = alto_px * 0.75 + 6   # px -> puntos


def _rotulo(ws, fila, col, texto, ancho_cols):
    """Banda de sección azul corporativa a lo ancho de `ancho_cols` columnas."""
    ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + ancho_cols - 1)
    _set(ws, fila, col, texto, font=F_CAB, fill=AZUL,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[fila].height = 18


def _campo(ws, fila, col, etiqueta, valor, max_cols=1):
    """Par etiqueta/valor: etiqueta en gris, valor en caja destacada.

    La caja se combina solo hasta que el texto quepa (como máximo `max_cols`
    columnas), para no dejar celdas de color con nada escrito. Sin ajuste de
    texto: Excel no autoajusta el alto de las combinadas y un valor largo (una
    dirección) se cortaría al envolver.
    """
    _set(ws, fila, col, etiqueta, font=F_NEG, fill=GRIS_CLARO, align=IZQ)
    texto = "" if valor is None else str(valor)

    n, ancho = 0, 0
    for k in range(max_cols):
        n += 1
        letra = get_column_letter(col + 1 + k)
        ancho += ws.column_dimensions[letra].width or 8.43
        if ancho >= len(texto) + 2:
            break
    if n > 1:
        ws.merge_cells(start_row=fila, start_column=col + 1, end_row=fila, end_column=col + n)
    _set(ws, fila, col + 1, texto, font=F_DATO, fill=AMARILLO, align=IZQ_1L)


def _tabla_consumo(ws, fila0, meses, totales, pcts, titulo_col="Mes"):
    """Tabla de consumo mensual por periodo (enero→diciembre) + TOTAL y %."""
    cabeceras = [titulo_col, "P1", "P2", "P3", "P4", "P5", "P6", "Total (kWh)"]
    for j, texto in enumerate(cabeceras, start=1):
        _set(ws, fila0, j, texto, font=F_CAB, fill=GRIS, align=CENTRO)

    f = fila0
    for m in meses:
        f += 1
        _set(ws, f, 1, m.get("abr") or m["nombre"], font=F_NEG, fill=GRIS_CLARO, align=IZQ)
        for j, v in enumerate(m["periodos"], start=2):
            _set(ws, f, j, float(v), fmt=FMT_KWH_ENT, align=CENTRO)
        _set(ws, f, 8, float(m["total"]), font=F_NEG, fmt=FMT_KWH_ENT, align=CENTRO)

    f += 1
    _set(ws, f, 1, "TOTAL", font=F_NEG, fill=GRIS_CLARO, align=IZQ)
    for j, v in enumerate(totales, start=2):
        _set(ws, f, j, float(v), font=F_NEG, fill=GRIS_CLARO, fmt=FMT_KWH_ENT, align=CENTRO)
    _set(ws, f, 8, float(sum(totales)), font=F_NEG, fill=GRIS_CLARO, fmt=FMT_KWH_ENT, align=CENTRO)

    f += 1
    _set(ws, f, 1, "%", font=F_NEG, fill=GRIS_CLARO, align=IZQ)
    for j, v in enumerate(pcts, start=2):
        _set(ws, f, j, float(v) / 100, font=F_NOR, fill=GRIS_CLARO, fmt=FMT_PCT0, align=CENTRO)
    _set(ws, f, 8, 1, font=F_NEG, fill=GRIS_CLARO, fmt=FMT_PCT0, align=CENTRO)
    return f


def _grafico_apilado(ws, fila_cab, n_meses, ancla):
    """Gráfico de barras apiladas por periodo, nativo de Excel."""
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.title = "Consumo mensual por periodo (kWh)"
    ch.height, ch.width = 9, 24
    # openpyxl crea los ejes con delete=True: sin esto el gráfico sale sin
    # etiquetas de meses ni escala en Y.
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.tickLblPos = "low"
    ch.y_axis.tickLblPos = "nextTo"
    ch.y_axis.majorGridlines = None
    # Leyenda debajo. openpyxl la crea con overlay=True, es decir, dibujada ENCIMA del
    # área del gráfico: por eso se montaba sobre la última barra. Con overlay=False
    # reserva su propia banda.
    ch.legend.position = "b"
    ch.legend.overlay = False

    datos = Reference(ws, min_col=2, max_col=7, min_row=fila_cab, max_row=fila_cab + n_meses)
    cats = Reference(ws, min_col=1, min_row=fila_cab + 1, max_row=fila_cab + n_meses)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)
    for i, serie in enumerate(ch.series):
        serie.graphicalProperties.solidFill = COLOR_PERIODO[i]
        # Filo blanco: en gama pastel los tramos apilados se confundirían.
        serie.graphicalProperties.line.solidFill = "FFFFFF"
        serie.graphicalProperties.line.width = 9525   # 0,75 pt en EMU
    ws.add_chart(ch, ancla)


def _hoja_suministro(wb, expediente, ficha, fecha):
    """Ficha de un punto de suministro: datos, potencias, consumo mensual y gráfico."""
    punto = ficha["punto"]
    ws = wb.create_sheet(title=_titulo_hoja(wb, f"Consumo {punto.cups[-6:]}"))
    ws.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDEFGH", (20, 14, 14, 14, 14, 18, 14, 15)):
        ws.column_dimensions[col].width = ancho
    _logo_centrado(ws, 8)

    _rotulo(ws, 3, 1, "CUESTIONARIO DE CONSUMO DE ENERGÍA ELÉCTRICA", 8)
    _campo(ws, 5, 1, "Empresa", expediente.cliente_razon_social, max_cols=4)
    _campo(ws, 5, 6, "CIF / NIF", expediente.cliente_cif, max_cols=2)
    _campo(ws, 6, 1, "Dirección", punto.direccion or expediente.cliente_direccion, max_cols=4)
    _campo(ws, 6, 6, "Fecha", fecha.strftime("%d/%m/%Y"), max_cols=2)
    _campo(ws, 7, 1, "CUPS", punto.cups, max_cols=4)
    _campo(ws, 7, 6, "Tarifa de acceso", punto.tarifa, max_cols=2)
    for f in (5, 6, 7):
        ws.row_dimensions[f].height = 17

    _rotulo(ws, 9, 1, "Datos del punto de suministro", 8)
    ws.merge_cells(start_row=10, start_column=1, end_row=11, end_column=1)
    _set(ws, 10, 1, "Potencia contratada (kW)", font=F_NEG, fill=GRIS_CLARO, align=IZQ)
    for j in range(6):
        _set(ws, 10, 2 + j, f"P{j + 1}", font=F_CAB, fill=GRIS, align=CENTRO)
    for j, p in enumerate(ficha["potencias"]):
        _set(ws, 11, 2 + j, float(p) if p is not None else "—",
             fill=AMARILLO, fmt=FMT_KW, align=CENTRO)
    _campo(ws, 12, 1, "Comercializadora", punto.comercializadora_actual or "—", max_cols=4)
    _campo(ws, 12, 6, "Fin de contrato",
           punto.fecha_fin_contrato.strftime("%d/%m/%Y") if punto.fecha_fin_contrato else "—",
           max_cols=2)
    ws.row_dimensions[12].height = 17

    _rotulo(ws, 14, 1, "Consumo realizado durante el periodo de un año", 8)
    fila_cab = 15
    fin = _tabla_consumo(ws, fila_cab, ficha["meses"], ficha["totales_periodo"], ficha["pct_periodo"])
    _grafico_apilado(ws, fila_cab, len(ficha["meses"]), f"A{fin + 3}")
    return ws


def _hoja_suministros_multiple(wb, expediente, resumen):
    """Un único cuadro con todos los CUPS: potencias y perfil de consumo anual."""
    ws = wb.create_sheet(title=_titulo_hoja(wb, "Suministros"))
    ws.sheet_view.showGridLines = False

    cab1 = [("Nº", 1), ("Razón social", 1), ("CIF", 1), ("Dirección de suministro", 1),
            ("CUPS", 1), ("Tarifa acceso", 1),
            ("Potencia máx. contratada (kW)", 6), ("Perfil de consumo anual (kWh)", 6),
            ("Consumo anual (kWh)", 1)]
    col = 1
    for texto, ancho in cab1:
        if ancho > 1:
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + ancho - 1)
            _set(ws, 3, col, texto, font=F_CAB, fill=VERDE_OSC, align=CENTRO)
            for k in range(ancho):
                _set(ws, 4, col + k, f"P{k + 1}", font=F_CAB, fill=GRIS, align=CENTRO)
        else:
            ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
            _set(ws, 3, col, texto, font=F_CAB, fill=VERDE_OSC, align=CENTRO)
        col += ancho
    n_cols = col - 1

    f = 4
    for i, ficha in enumerate(resumen["fichas"], start=1):
        f += 1
        p = ficha["punto"]
        _set(ws, f, 1, i, font=F_NEG, fill=AMARILLO, align=CENTRO)
        _set(ws, f, 2, p.titular or expediente.cliente_razon_social, align=IZQ)
        _set(ws, f, 3, p.cif_titular or expediente.cliente_cif, align=CENTRO)
        _set(ws, f, 4, p.direccion or "—", align=IZQ)
        _set(ws, f, 5, p.cups, align=CENTRO)
        _set(ws, f, 6, p.tarifa, align=CENTRO)
        for j, pot in enumerate(ficha["potencias"]):
            _set(ws, f, 7 + j, float(pot) if pot is not None else "—", fmt=FMT_KW, align=CENTRO)
        for j, tot in enumerate(ficha["totales_periodo"]):
            _set(ws, f, 13 + j, float(tot), fmt=FMT_KWH_ENT, align=CENTRO)
        _set(ws, f, 19, float(ficha["total_anual"]), font=F_NEG, fmt=FMT_KWH_ENT, align=CENTRO)

    f += 1
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=18)
    _set(ws, f, 1, "TOTAL", font=F_NEG, fill=GRIS_CLARO, align=Alignment(horizontal="right"))
    _set(ws, f, 19, float(resumen["total_global"]), font=F_NEG, fill=GRIS_CLARO,
         fmt=FMT_KWH_ENT, align=CENTRO)

    for col, ancho in (("A", 5), ("B", 30), ("C", 12), ("D", 38), ("E", 24), ("F", 11), ("S", 16)):
        ws.column_dimensions[col].width = ancho
    for j in range(7, 19):
        ws.column_dimensions[ws.cell(row=4, column=j).column_letter].width = 11
    ws.freeze_panes = "G5"
    return ws


def _hoja_consumo_agregado(wb, resumen):
    """Consumo mensual sumado de todos los suministros, en su propia hoja.

    Va aparte del cuadro de suministros porque ese necesita 19 columnas estrechas
    y aquí hacen falta 8 anchas: compartir hoja descuadraría una de las dos.
    """
    ws = wb.create_sheet(title=_titulo_hoja(wb, "Consumo agregado"))
    ws.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDEFGH", (14, 14, 14, 14, 14, 14, 14, 15)):
        ws.column_dimensions[col].width = ancho

    _rotulo(ws, 2, 1, "Consumo mensual agregado de todos los suministros (kWh)", 8)
    fila_cab = 3
    fin = _tabla_consumo(ws, fila_cab, resumen["agg_meses"],
                         resumen["totales_periodo"], resumen["pct_periodo"])
    _grafico_apilado(ws, fila_cab, len(resumen["agg_meses"]), f"A{fin + 3}")
    return ws


def _titulo_hoja(wb, base):
    """Nombre de hoja válido y único (Excel: 31 caracteres, sin duplicados)."""
    base = base[:31]
    titulo, i = base, 2
    existentes = {n.lower() for n in wb.sheetnames}
    while titulo.lower() in existentes:
        titulo = f"{base[:28]}_{i}"
        i += 1
    return titulo


def _hojas_datos(wb, expediente):
    """Añade al libro los datos de partida del estudio, ya formateados.

    Se generan desde la base de datos, no se copian del Excel subido: la plantilla
    trae hojas de instrucciones y decenas de filas y bloques vacíos que no aportan
    nada al informe.
    """
    resumen = resumen_consumo(expediente)
    if not resumen:
        return
    fecha = timezone.localdate()
    if resumen["multiple"]:
        _hoja_suministros_multiple(wb, expediente, resumen)
        _hoja_consumo_agregado(wb, resumen)
        for ficha in resumen["fichas"]:
            _hoja_suministro(wb, expediente, ficha, fecha)
    else:
        _hoja_suministro(wb, expediente, resumen["fichas"][0], fecha)


def comparativo_a_pdf(expediente, comparativo, detalle_cups, consumo=None):
    mejor = next((c for c in comparativo["columnas"] if c.get("mejor")), None)
    if mejor and mejor["ahorro"] is not None:
        mejor["ahorro_abs"] = abs(mejor["ahorro"])

    # El gráfico se dibuja como PNG y se incrusta: xhtml2pdf no sabe dibujar barras.
    grafico = None
    if consumo and consumo["meses"]:
        png = grafico_consumo_apilado(consumo["meses"])
        if png:
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp.write(png)
            tmp.close()
            grafico = tmp.name
    try:
        html = render_to_string("estudios/informe_pdf.html", {
            "expediente": expediente,
            "comparativo": comparativo,
            "detalle_cups": detalle_cups,
            "consumo": consumo,
            "grafico": grafico,
            "mejor": mejor,
            "logo": str(settings.BASE_DIR / LOGO),
        })
        buffer = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=buffer, encoding="utf-8")
        return buffer.getvalue()
    finally:
        if grafico:
            os.unlink(grafico)
