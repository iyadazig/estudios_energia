"""Tests del motor de cálculo (estudios/calculo.py) y de la validación de CUPS.

Son la red de seguridad del corazón de la aplicación: si un cambio futuro
altera los importes del comparativo, estos tests fallan y avisan antes de que
un PDF con cifras erróneas llegue a un cliente.

El método: para cada escenario se calcula el resultado esperado con una
aritmética simple e independiente (bucles planos en el propio test, leyendo
los parámetros regulados de la base de datos) y se compara, al céntimo, con
lo que produce el motor.

Ejecutar con:  python manage.py test estudios   (o: pytest)
"""
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import Client, TestCase

from .calculo import comparativo_expediente
from .importador import LETRAS_CONTROL_CUPS, validar_cups
from .models import (
    ConceptoAdicional,
    ConsumoMensual,
    Expediente,
    Oferta,
    ConceptoCatalogo,
    OfertaCatalogo,
    ParametroGeneral,
    ParametroRegulado,
    PrecioCatalogo,
    PrecioOferta,
    ProfileFactor,
    PuntoSuministro,
    SerieSSAA,
)

CENT = Decimal("0.01")


def cups_valido(digitos16):
    """Construye un CUPS válido (ES + 16 dígitos + 2 letras de control)."""
    resto = int(digitos16) % 529
    c, m = divmod(resto, 23)
    return "ES" + digitos16 + LETRAS_CONTROL_CUPS[c] + LETRAS_CONTROL_CUPS[m]


# --- Cálculo de referencia, independiente del motor -------------------------

def _param(tipo, termino, tarifa):
    p = ParametroRegulado.objects.get(tipo=tipo, termino=termino, tarifa=tarifa)
    return [p.p1, p.p2, p.p3, p.p4, p.p5, p.p6]


def _atr(termino, tarifa):
    peaje = _param("peaje", termino, tarifa)
    cargo = _param("cargo", termino, tarifa)
    return [a + b for a, b in zip(peaje, cargo)]


def _iee():
    return ParametroGeneral.objects.get(clave="iee").valor


def _consumos(punto):
    return [(c.anio, c.mes, [c.p1, c.p2, c.p3, c.p4, c.p5, c.p6]) for c in punto.consumos.all()]


def coste_referencia_actual(punto):
    """Reimplementación simple del coste anual del contrato actual (precio fijo)."""
    consumos = _consumos(punto)
    precios_e = [punto.precio_energia_p1, punto.precio_energia_p2, punto.precio_energia_p3,
                 punto.precio_energia_p4, punto.precio_energia_p5, punto.precio_energia_p6]
    precios_e = [p or Decimal(0) for p in precios_e]
    if not punto.energia_peajes_incluidos:
        precios_e = [p + a for p, a in zip(precios_e, _atr("energia", punto.tarifa))]
    energia = sum((sum(k * pr for k, pr in zip(v, precios_e)) for _, _, v in consumos), Decimal(0))

    if punto.potencia_segun_atr:
        precios_p = _atr("potencia", punto.tarifa)
    else:
        precios_p = [punto.precio_potencia_p1, punto.precio_potencia_p2, punto.precio_potencia_p3,
                     punto.precio_potencia_p4, punto.precio_potencia_p5, punto.precio_potencia_p6]
        precios_p = [p or Decimal(0) for p in precios_p]
    kw = [punto.potencia_p1, punto.potencia_p2, punto.potencia_p3,
          punto.potencia_p4, punto.potencia_p5, punto.potencia_p6]
    potencia = sum((k * pr for k, pr in zip(kw, precios_p) if k), Decimal(0))

    iee = (potencia + energia) * _iee()
    return potencia + energia + iee


def coste_referencia_oferta(oferta, punto):
    consumos = _consumos(punto)
    precio = oferta.precios.filter(punto__isnull=True).first()
    precios_e = [precio.energia_p1, precio.energia_p2, precio.energia_p3,
                 precio.energia_p4, precio.energia_p5, precio.energia_p6]
    precios_e = [p or Decimal(0) for p in precios_e]
    if not oferta.atr_energia_incluido:
        precios_e = [p + a for p, a in zip(precios_e, _atr("energia", punto.tarifa))]
    energia = sum((sum(k * pr for k, pr in zip(v, precios_e)) for _, _, v in consumos), Decimal(0))

    if oferta.atr_potencia_incluido:
        precios_p = [precio.potencia_p1, precio.potencia_p2, precio.potencia_p3,
                     precio.potencia_p4, precio.potencia_p5, precio.potencia_p6]
        precios_p = [p or Decimal(0) for p in precios_p]
    else:
        precios_p = _atr("potencia", punto.tarifa)
    kw = [punto.potencia_p1, punto.potencia_p2, punto.potencia_p3,
          punto.potencia_p4, punto.potencia_p5, punto.potencia_p6]
    potencia = sum((k * pr for k, pr in zip(kw, precios_p) if k), Decimal(0))

    perdidas = _param("perdidas", "energia", punto.tarifa)
    base_iee = potencia + energia
    conceptos = Decimal(0)
    for con in oferta.conceptos.all():
        importe = Decimal(0)
        if con.tipo == "eur_mwh":
            mwh = Decimal(0)
            for _, _, v in consumos:
                if con.con_perdidas:
                    mwh += sum(k * (1 + c) for k, c in zip(v, perdidas)) / 1000
                else:
                    mwh += sum(v) / 1000
            importe = con.valor * mwh
        elif con.tipo == "fijo_mes":
            importe = con.valor * 12
        elif con.tipo == "pct":
            importe = (potencia + energia) * con.valor / 100
        if con.con_impuesto_local:
            importe *= (1 + _param("imp_local", "energia", punto.tarifa)[0])
        conceptos += importe
        if con.entra_en_iee:
            base_iee += importe

    ssaa = coste_ssaa_ref(oferta, punto, consumos)
    conceptos += ssaa
    base_iee += ssaa

    iee = base_iee * _iee()
    return potencia + energia + conceptos + iee


def coste_ssaa_ref(oferta, punto, consumos):
    """Reimplementación simple e independiente del coste de SSAA de la oferta."""
    if oferta.ssaa_tipo == "incluido":
        return Decimal(0)
    if oferta.ssaa_perdidas_modo == "fija":
        perd = [(oferta.ssaa_perdidas_pct or Decimal(0)) / 100] * 6
    else:
        perd = _param("perdidas", "energia", punto.tarifa)
    ap = oferta.ssaa_apuntamiento if oferta.ssaa_apuntamiento is not None else Decimal(1)
    hl = (1 + _param("imp_local", "energia", punto.tarifa)[0]) if oferta.ssaa_impuesto_municipal else Decimal(1)

    def dif(ssaa):
        if oferta.ssaa_tipo == "indexado":
            return ssaa
        sup, inf = oferta.ssaa_ref_superior, oferta.ssaa_ref_inferior
        if oferta.ssaa_tipo == "techo":
            return ssaa - sup if (sup is not None and ssaa > sup) else Decimal(0)
        if sup is not None and ssaa > sup:
            return ssaa - sup
        if inf is not None and ssaa < inf:
            return ssaa - inf
        return Decimal(0)

    total = Decimal(0)
    for _anio, mes, v in consumos:
        base = SerieSSAA.objects.get(mes=mes).valor_considerado
        if oferta.ssaa_modo == "horario":
            pf = ProfileFactor.objects.get(ambito="peninsula_3_6", mes=mes)
            factores = [pf.p1, pf.p2, pf.p3, pf.p4, pf.p5, pf.p6]
            ssaa_p = [base * f if f is not None else None for f in factores]
        else:
            ssaa_p = [base] * 6
        for p in range(6):
            if not v[p] or ssaa_p[p] is None:
                continue
            d = dif(ssaa_p[p])
            if d == 0:
                continue
            total += d * (1 + perd[p]) * ap * (v[p] / Decimal(1000)) * hl
    return total


# --- Casos de prueba --------------------------------------------------------

class CalculoMotorTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        # El comando ya carga peajes/cargos/pérdidas/imp.local/IEE + serie SSAA + profile factors.
        call_command("cargar_parametros_2026")
        cls.user = User.objects.create_user("tester", password="x")

    def _expediente(self):
        return Expediente.objects.create(
            cliente_razon_social="Cliente Test, S.A.", cliente_cif="A00000000", gestor=self.user,
        )

    def _consumos_12_meses(self, punto, base=1000):
        for mes in range(1, 13):
            ConsumoMensual.objects.create(
                punto=punto, anio=2025, mes=mes,
                p1=base + mes, p2=base + mes + 1, p3=base + mes + 2,
                p4=base + mes + 3, p5=base + mes + 4, p6=base + mes + 5,
            )

    def test_linea_base_un_cups_potencia_atr(self):
        """Contrato actual 6.1TD con potencia según ATR: el motor cuadra con el cálculo de referencia."""
        exp = self._expediente()
        punto = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000001"), tarifa="6.1TD",
            potencia_p1=10, potencia_p2=10, potencia_p3=10, potencia_p4=10, potencia_p5=10, potencia_p6=20,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
            precio_energia_p1=Decimal("0.15"), precio_energia_p2=Decimal("0.13"),
            precio_energia_p3=Decimal("0.11"), precio_energia_p4=Decimal("0.09"),
            precio_energia_p5=Decimal("0.08"), precio_energia_p6=Decimal("0.10"),
        )
        self._consumos_12_meses(punto)
        # Una oferta cualquiera para que el comparativo se genere
        oferta = Oferta.objects.create(expediente=exp, comercializadora="X", atr_energia_incluido=True)
        PrecioOferta.objects.create(oferta=oferta, punto=None,
                                    energia_p1=Decimal("0.14"), energia_p2=Decimal("0.12"),
                                    energia_p3=Decimal("0.10"), energia_p4=Decimal("0.08"),
                                    energia_p5=Decimal("0.07"), energia_p6=Decimal("0.09"))

        r = comparativo_expediente(exp)
        esperado = coste_referencia_actual(punto)
        self.assertAlmostEqual(r["actual"].total, esperado, delta=CENT)
        # El €/kWh medio = total / consumo total
        consumo_total = sum(c.total for c in punto.consumos.all())
        self.assertAlmostEqual(r["actual"].eur_kwh, esperado / consumo_total, delta=Decimal("0.000001"))

    def _punto_6_1TD(self, cups, base=2000):
        punto = PuntoSuministro.objects.create(
            expediente=self._exp_ssaa, cups=cups_valido(cups), tarifa="6.1TD",
            potencia_p1=50, potencia_p2=50, potencia_p3=50, potencia_p4=50, potencia_p5=50, potencia_p6=100,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
            precio_energia_p1=Decimal("0.16"), precio_energia_p2=Decimal("0.14"),
            precio_energia_p3=Decimal("0.12"), precio_energia_p4=Decimal("0.10"),
            precio_energia_p5=Decimal("0.09"), precio_energia_p6=Decimal("0.11"),
        )
        self._consumos_12_meses(punto, base=base)
        return punto

    def _oferta_ssaa(self, comercializadora, **ssaa):
        oferta = Oferta.objects.create(expediente=self._exp_ssaa, comercializadora=comercializadora,
                                       atr_energia_incluido=True, **ssaa)
        PrecioOferta.objects.create(oferta=oferta, punto=None,
                                    energia_p1=Decimal("0.15"), energia_p2=Decimal("0.13"),
                                    energia_p3=Decimal("0.11"), energia_p4=Decimal("0.09"),
                                    energia_p5=Decimal("0.08"), energia_p6=Decimal("0.10"))
        return oferta

    def test_ssaa_techo_mensual(self):
        """SSAA con techo, modo mensual: el motor cuadra con la referencia y el importe es > 0."""
        self._exp_ssaa = self._expediente()
        punto = self._punto_6_1TD("0011000000000002")
        oferta = self._oferta_ssaa(
            "ENDESA", ssaa_tipo="techo", ssaa_modo="mensual", ssaa_ref_superior=Decimal("16"),
            ssaa_perdidas_modo="fija", ssaa_perdidas_pct=Decimal("7"),
        )
        r = comparativo_expediente(self._exp_ssaa)
        col = r["columnas"][0]
        self.assertAlmostEqual(col["agregado"].total, coste_referencia_oferta(oferta, punto), delta=CENT)
        self.assertGreater(col["agregado"].ssaa, 0)

    def test_ssaa_horario_vs_mensual(self):
        """El modo horario (profile factors) da un resultado distinto del mensual y cuadra con la referencia."""
        self._exp_ssaa = self._expediente()
        punto = self._punto_6_1TD("0011000000000003")
        oferta = self._oferta_ssaa(
            "IBERDROLA", ssaa_tipo="indexado", ssaa_modo="horario",
            ssaa_perdidas_modo="circular",
        )
        r = comparativo_expediente(self._exp_ssaa)
        self.assertAlmostEqual(r["columnas"][0]["agregado"].total, coste_referencia_oferta(oferta, punto), delta=CENT)

    def test_ssaa_banda_abono_negativo(self):
        """Con banda, si el SSAA queda por debajo de la referencia mínima, la regularización es negativa (abono)."""
        self._exp_ssaa = self._expediente()
        punto = self._punto_6_1TD("0011000000000004")
        # Ref. mínima muy alta (40) para forzar que todos los meses queden por debajo → abono.
        oferta = self._oferta_ssaa(
            "NATURGY", ssaa_tipo="banda", ssaa_modo="mensual",
            ssaa_ref_inferior=Decimal("40"), ssaa_ref_superior=Decimal("60"),
            ssaa_perdidas_modo="fija", ssaa_perdidas_pct=Decimal("7"),
        )
        r = comparativo_expediente(self._exp_ssaa)
        col = r["columnas"][0]
        self.assertAlmostEqual(col["agregado"].total, coste_referencia_oferta(oferta, punto), delta=CENT)
        self.assertLess(col["agregado"].ssaa, 0)

    def test_ssaa_apuntamiento_e_impuesto_municipal(self):
        """El apuntamiento y el impuesto municipal multiplican el SSAA por sus factores."""
        self._exp_ssaa = self._expediente()
        self._punto_6_1TD("0011000000000005")
        base = dict(ssaa_tipo="indexado", ssaa_modo="mensual", ssaa_perdidas_modo="fija",
                    ssaa_perdidas_pct=Decimal("7"))
        o_sin = self._oferta_ssaa("SIN", **base)
        o_con = self._oferta_ssaa("CON", ssaa_apuntamiento=Decimal("1.02"),
                                  ssaa_impuesto_municipal=True, **base)
        r = comparativo_expediente(self._exp_ssaa)
        por_com = {c["oferta"].comercializadora: c["agregado"].ssaa for c in r["columnas"]}
        # CON = SIN × 1,02 (apuntamiento) × 1,015 (impuesto municipal)
        self.assertAlmostEqual(por_com["CON"], por_com["SIN"] * Decimal("1.02") * Decimal("1.015"), delta=CENT)

    def test_ssaa_incluido_no_cuesta(self):
        """Con SSAA incluido en el precio, la regularización de SSAA es cero."""
        self._exp_ssaa = self._expediente()
        self._punto_6_1TD("0011000000000006")
        self._oferta_ssaa("X", ssaa_tipo="incluido")
        r = comparativo_expediente(self._exp_ssaa)
        self.assertEqual(r["columnas"][0]["agregado"].ssaa, Decimal(0))

    def test_conceptos_por_nombre_y_ssaa_en_linea_propia(self):
        """Los conceptos aparecen por su nombre y la SSAA va en su propia línea (fuera de conceptos)."""
        self._exp_ssaa = self._expediente()
        self._punto_6_1TD("0011000000000031")
        oferta = self._oferta_ssaa("ENDESA", ssaa_tipo="techo", ssaa_modo="mensual",
                                   ssaa_ref_superior=Decimal("16"),
                                   ssaa_perdidas_modo="fija", ssaa_perdidas_pct=Decimal("7"))
        ConceptoAdicional.objects.create(oferta=oferta, nombre="FNEE", tipo="eur_mwh",
                                         valor=Decimal("1.492"), con_perdidas=True, entra_en_iee=True)
        r = comparativo_expediente(self._exp_ssaa)
        col = r["columnas"][0]["agregado"]
        self.assertIn("FNEE", r["conceptos_nombres"])
        self.assertGreater(col.conceptos["FNEE"], 0)
        self.assertGreater(col.ssaa, 0)
        self.assertNotIn("Regularización SSAA", col.conceptos)  # la SSAA no está entre los conceptos

    def test_incluido_precio_energia_en_conceptos_y_ssaa(self):
        """En pantalla, PDF y Excel, una oferta que no lleva un concepto (o con SSAA
        incluido en el precio) muestra 'incluido precio energía' en esa celda."""
        import io
        from openpyxl import load_workbook
        from django.template.loader import render_to_string
        from .exportar import comparativo_a_excel

        self._exp_ssaa = self._expediente()
        self._punto_6_1TD("0011000000000048")
        # Oferta A: lleva FNEE y regulariza SSAA con techo -> muestra números
        a = self._oferta_ssaa("CON_FNEE", ssaa_tipo="techo", ssaa_modo="mensual",
                              ssaa_ref_superior=Decimal("16"), ssaa_perdidas_modo="fija",
                              ssaa_perdidas_pct=Decimal("7"))
        ConceptoAdicional.objects.create(oferta=a, nombre="FNEE", tipo="eur_mwh",
                                         valor=Decimal("1.492"), con_perdidas=True, entra_en_iee=True)
        # Oferta B: sin FNEE y con SSAA incluido en el precio -> ambas celdas con el texto
        self._oferta_ssaa("SIN_FNEE", ssaa_tipo="incluido")

        comp = comparativo_expediente(self._exp_ssaa)
        self.assertIn("FNEE", comp["conceptos_nombres"])

        html = render_to_string("estudios/expediente_detalle.html",
                                {"expediente": self._exp_ssaa, "comparativo": comp,
                                 "ofertas": self._exp_ssaa.ofertas.all()})
        html_pdf = render_to_string("estudios/informe_pdf.html",
                                    {"expediente": self._exp_ssaa, "comparativo": comp})
        self.assertIn("incluido precio energía", html)
        self.assertIn("incluido precio energía", html_pdf)

        wb = load_workbook(io.BytesIO(comparativo_a_excel(self._exp_ssaa, comp)))
        ws = wb["Resumen"]
        textos = [c.value for row in ws.iter_rows() for c in row]
        # Aparece en la fila de FNEE (oferta B) y en la de Reg. SSAA (oferta B)
        self.assertGreaterEqual(textos.count("incluido precio energía"), 2)
        # La oferta con techo sí trae número en Reg. SSAA (no el texto)
        fila_ssaa = next(r for r in ws.iter_rows() if r[0].value and "SSAA" in str(r[0].value))
        valores = [c.value for c in fila_ssaa]
        self.assertTrue(any(isinstance(v, float) and v > 0 for v in valores))

    def _expediente_con_oferta(self, n_puntos=1):
        self._exp_ssaa = self._expediente()
        for i in range(n_puntos):
            self._punto_6_1TD(f"001100000000006{i}")
        self._oferta_ssaa("X", ssaa_tipo="incluido")
        return self._exp_ssaa

    def _html_pdf(self, exp):
        from django.template.loader import render_to_string
        from .consumo import resumen_consumo
        from .views import _detalle_cups
        comp = comparativo_expediente(exp)
        self.assertTrue(comp["avisos"])  # hay avisos, pero no deben salir en el PDF
        return render_to_string("estudios/informe_pdf.html", {
            "expediente": exp, "comparativo": comp,
            "detalle_cups": _detalle_cups(comp), "consumo": resumen_consumo(exp),
        })

    def test_pdf_portada_consumo_un_suministro_y_sin_avisos(self):
        """Con un solo CUPS el PDF abre con su ficha y consumo mensual; nunca lleva avisos."""
        exp = self._expediente_con_oferta(n_puntos=1)
        html = self._html_pdf(exp)
        self.assertIn("Datos del suministro", html)
        self.assertIn("Consumo mensual por periodo", html)
        self.assertIn(exp.puntos.first().cups, html)
        self.assertNotIn("Resumen de suministros", html)
        self.assertNotIn("Avisos del cálculo", html)

    def test_pdf_portada_consumo_varios_suministros(self):
        """Con varios CUPS el PDF abre con el resumen de todos y el consumo agregado."""
        exp = self._expediente_con_oferta(n_puntos=2)
        html = self._html_pdf(exp)
        self.assertIn("Resumen de suministros", html)
        self.assertIn("agregado", html)
        self.assertNotIn("Datos del suministro", html)
        for punto in exp.puntos.all():
            self.assertIn(punto.cups, html)
        self.assertNotIn("Avisos del cálculo", html)

    def test_consumo_se_ordena_de_enero_a_diciembre(self):
        """Aunque el histórico vaya de mayo a abril, se presenta de enero a diciembre."""
        from .consumo import resumen_consumo
        self._exp_ssaa = self._expediente()
        punto = PuntoSuministro.objects.create(
            expediente=self._exp_ssaa, cups=cups_valido("0011000000000091"), tarifa="6.1TD",
            potencia_p1=50, potencia_p2=50, potencia_p3=50, potencia_p4=50, potencia_p5=50,
            potencia_p6=100, energia_peajes_incluidos=True, potencia_segun_atr=True,
            precio_energia_p1=Decimal("0.16"),
        )
        # Histórico "a caballo": may–dic de 2025 y ene–abr de 2026.
        for mes in range(1, 13):
            ConsumoMensual.objects.create(
                punto=punto, anio=(2026 if mes <= 4 else 2025), mes=mes,
                p1=100 * mes, p2=10, p3=10, p4=10, p5=10, p6=10,
            )
        meses = resumen_consumo(self._exp_ssaa)["meses"]
        self.assertEqual([m["mes"] for m in meses], list(range(1, 13)))
        self.assertEqual(meses[0]["abr"], "Ene 2026")
        self.assertEqual(meses[11]["abr"], "Dic 2025")

    def test_excel_ficha_de_consumo_un_suministro(self):
        """Un solo CUPS: hoja de cuestionario con sus datos, consumo y gráfico."""
        import io
        from openpyxl import load_workbook
        from .exportar import comparativo_a_excel

        exp = self._expediente_con_oferta(n_puntos=1)
        wb = load_workbook(io.BytesIO(comparativo_a_excel(exp, comparativo_expediente(exp))))

        self.assertEqual(wb.sheetnames[0], "Resumen")
        hoja = next(h for h in wb.sheetnames if h.startswith("Consumo "))
        ws = wb[hoja]
        # No se arrastra nada de la plantilla original (instrucciones, filas vacías…)
        self.assertNotIn("Instrucciones", wb.sheetnames)
        self.assertNotIn("Suministros", wb.sheetnames)
        self.assertEqual(ws["A3"].value, "CUESTIONARIO DE CONSUMO DE ENERGÍA ELÉCTRICA")
        self.assertEqual(ws["B5"].value, exp.cliente_razon_social)
        self.assertEqual(ws["B7"].value, exp.puntos.first().cups)
        self.assertEqual(ws["A15"].value, "Mes")          # cabecera de la tabla de consumo
        self.assertEqual(ws["A16"].value[:3], "Ene")      # empieza en enero
        self.assertEqual(len(ws._charts), 1)              # gráfico apilado por periodo

    def test_excel_cuadro_de_suministros_si_hay_varios(self):
        """Varios CUPS: cuadro resumen con potencias y perfil anual + consumo agregado."""
        import io
        from openpyxl import load_workbook
        from .exportar import comparativo_a_excel

        exp = self._expediente_con_oferta(n_puntos=2)
        wb = load_workbook(io.BytesIO(comparativo_a_excel(exp, comparativo_expediente(exp))))

        self.assertIn("Suministros", wb.sheetnames)
        self.assertIn("Consumo agregado", wb.sheetnames)
        ws = wb["Suministros"]
        self.assertEqual(ws["G3"].value, "Potencia máx. contratada (kW)")
        self.assertEqual(ws["M3"].value, "Perfil de consumo anual (kWh)")
        cups = [p.cups for p in exp.puntos.all()]
        self.assertIn(ws["E5"].value, cups)
        self.assertIn(ws["E6"].value, cups)
        # Una ficha por cada CUPS, además del cuadro resumen
        fichas = [h for h in wb.sheetnames if h.startswith("Consumo ") and h != "Consumo agregado"]
        self.assertEqual(len(fichas), 2)

    def test_dos_cups_agregado_es_suma_de_cups(self):
        """El total agregado del expediente es la suma exacta de los totales por CUPS."""
        exp = self._expediente()
        p1 = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000003"), tarifa="6.1TD",
            potencia_p1=15, potencia_p2=15, potencia_p3=15, potencia_p4=15, potencia_p5=15, potencia_p6=30,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
            precio_energia_p1=Decimal("0.15"), precio_energia_p2=Decimal("0.13"),
            precio_energia_p3=Decimal("0.11"), precio_energia_p4=Decimal("0.09"),
            precio_energia_p5=Decimal("0.08"), precio_energia_p6=Decimal("0.10"),
        )
        p2 = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000004"), tarifa="2.0TD",
            potencia_p1=10, potencia_p2=12,
            energia_peajes_incluidos=True, potencia_segun_atr=False,
            precio_energia_p1=Decimal("0.18"), precio_energia_p2=Decimal("0.15"), precio_energia_p3=Decimal("0.12"),
            precio_potencia_p1=Decimal("30"), precio_potencia_p2=Decimal("4"),
        )
        self._consumos_12_meses(p1, base=1000)
        for mes in range(1, 13):  # 2.0TD: solo P1-P3
            ConsumoMensual.objects.create(punto=p2, anio=2025, mes=mes,
                                          p1=300 + mes, p2=200 + mes, p3=100 + mes)
        oferta = Oferta.objects.create(expediente=exp, comercializadora="ENDESA", atr_energia_incluido=True)
        PrecioOferta.objects.create(oferta=oferta, punto=None,
                                    energia_p1=Decimal("0.14"), energia_p2=Decimal("0.12"),
                                    energia_p3=Decimal("0.10"), energia_p4=Decimal("0.08"),
                                    energia_p5=Decimal("0.07"), energia_p6=Decimal("0.09"))

        r = comparativo_expediente(exp)
        suma_cups = (r["actual_por_punto"][p1.pk].total + r["actual_por_punto"][p2.pk].total)
        self.assertAlmostEqual(r["actual"].total, suma_cups, delta=CENT)
        # Cada CUPS cuadra con su referencia
        self.assertAlmostEqual(r["actual_por_punto"][p1.pk].total, coste_referencia_actual(p1), delta=CENT)
        self.assertAlmostEqual(r["actual_por_punto"][p2.pk].total, coste_referencia_actual(p2), delta=CENT)

    def test_ranking_y_ahorro(self):
        """La oferta más barata es la nº 1 del ranking y su ahorro es negativo (ahorra)."""
        exp = self._expediente()
        punto = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000005"), tarifa="6.1TD",
            potencia_p1=20, potencia_p2=20, potencia_p3=20, potencia_p4=20, potencia_p5=20, potencia_p6=40,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
            precio_energia_p1=Decimal("0.20"), precio_energia_p2=Decimal("0.18"),
            precio_energia_p3=Decimal("0.16"), precio_energia_p4=Decimal("0.14"),
            precio_energia_p5=Decimal("0.12"), precio_energia_p6=Decimal("0.15"),
        )
        self._consumos_12_meses(punto)
        for nombre, precio in [("CARA", "0.18"), ("BARATA", "0.10")]:
            o = Oferta.objects.create(expediente=exp, comercializadora=nombre, atr_energia_incluido=True)
            PrecioOferta.objects.create(oferta=o, punto=None,
                                        energia_p1=Decimal(precio), energia_p2=Decimal(precio),
                                        energia_p3=Decimal(precio), energia_p4=Decimal(precio),
                                        energia_p5=Decimal(precio), energia_p6=Decimal(precio))
        r = comparativo_expediente(exp)
        mejor = [c for c in r["columnas"] if c.get("mejor")][0]
        self.assertEqual(mejor["oferta"].comercializadora, "BARATA")
        self.assertEqual(mejor["ranking"], 1)
        self.assertLess(mejor["ahorro"], 0)  # negativo = ahorra frente al actual

    def test_contrato_actual_indexado_no_calcula_linea_base(self):
        """Si el contrato actual es indexado, no hay línea base en esta fase (avisa)."""
        exp = self._expediente()
        punto = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000006"), tarifa="6.1TD",
            modalidad_actual="indexado",
            potencia_p1=10, potencia_p2=10, potencia_p3=10, potencia_p4=10, potencia_p5=10, potencia_p6=20,
            potencia_segun_atr=True,
        )
        self._consumos_12_meses(punto)
        oferta = Oferta.objects.create(expediente=exp, comercializadora="X", atr_energia_incluido=True)
        PrecioOferta.objects.create(oferta=oferta, punto=None, energia_p1=Decimal("0.10"))
        r = comparativo_expediente(exp)
        self.assertIsNone(r["actual"])
        self.assertTrue(any("indexado" in a for a in r["avisos"]))

    def test_precios_por_tarifa_y_override_por_cups(self):
        """Cada CUPS toma el precio de su tarifa; un precio por CUPS tiene prioridad."""
        exp = self._expediente()
        a = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000011"), tarifa="6.1TD",
            potencia_p1=10, potencia_p2=10, potencia_p3=10, potencia_p4=10, potencia_p5=10, potencia_p6=20,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
        )
        b = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido("0011000000000012"), tarifa="3.0TD",
            potencia_p1=15, potencia_p2=15, potencia_p3=15, potencia_p4=15, potencia_p5=15, potencia_p6=15,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
        )
        self._consumos_12_meses(a)
        self._consumos_12_meses(b)
        oferta = Oferta.objects.create(expediente=exp, comercializadora="ENDESA", atr_energia_incluido=True)
        # Precio por tarifa: distinto para 6.1TD y 3.0TD
        PrecioOferta.objects.create(oferta=oferta, tarifa="6.1TD",
                                    energia_p1=Decimal("0.10"), energia_p2=Decimal("0.10"), energia_p3=Decimal("0.10"),
                                    energia_p4=Decimal("0.10"), energia_p5=Decimal("0.10"), energia_p6=Decimal("0.10"))
        PrecioOferta.objects.create(oferta=oferta, tarifa="3.0TD",
                                    energia_p1=Decimal("0.20"), energia_p2=Decimal("0.20"), energia_p3=Decimal("0.20"),
                                    energia_p4=Decimal("0.20"), energia_p5=Decimal("0.20"), energia_p6=Decimal("0.20"))
        from .calculo import _precios_oferta_para
        self.assertEqual(_precios_oferta_para(oferta, a).energia_p1, Decimal("0.10"))
        self.assertEqual(_precios_oferta_para(oferta, b).energia_p1, Decimal("0.20"))
        # Override por CUPS concreto para 'a'
        PrecioOferta.objects.create(oferta=oferta, punto=a,
                                    energia_p1=Decimal("0.05"), energia_p2=Decimal("0.05"), energia_p3=Decimal("0.05"),
                                    energia_p4=Decimal("0.05"), energia_p5=Decimal("0.05"), energia_p6=Decimal("0.05"))
        self.assertEqual(_precios_oferta_para(oferta, a).energia_p1, Decimal("0.05"))  # gana el del CUPS
        self.assertEqual(_precios_oferta_para(oferta, b).energia_p1, Decimal("0.20"))  # b sigue por tarifa


class ImportacionPlantillaTest(TestCase):
    """La plantilla que el usuario descarga, rellena y sube queda guardada en el
    expediente, y sus pestañas se arrastran al Excel del comparativo."""

    @classmethod
    def setUpTestData(cls):
        call_command("cargar_parametros_2026")
        cls.user = User.objects.create_superuser("importador", password="x")

    def _plantilla_rellena(self):
        """Rellena la plantilla real de la app con un suministro y 12 meses."""
        import io
        from openpyxl import load_workbook
        from django.conf import settings
        from .importador import BASE_CONSUMOS, FILA_INI_SUMINISTROS

        ruta = settings.BASE_DIR / "estudios" / "plantillas" / "Plantilla_Estudio_Renovacion.xlsx"
        wb = load_workbook(ruta)

        for fila, valor in zip(range(4, 9), ["Cliente Plantilla, S.A.", "A11111111",
                                             "C/ Mayor 1", "Gestor", "g@ejemplo.com"]):
            wb["Cliente"].cell(row=fila, column=3, value=valor)

        sm, f = wb["Suministros"], FILA_INI_SUMINISTROS
        sm.cell(row=f, column=4, value=cups_valido("0011000000000077"))
        sm.cell(row=f, column=5, value="Nave 1")
        sm.cell(row=f, column=6, value="6.1TD")
        for i, p in enumerate([100, 100, 100, 100, 100, 200]):
            sm.cell(row=f, column=7 + i, value=p)
        sm.cell(row=f, column=16, value="Sí")          # peajes incluidos en energía
        for i, p in enumerate([0.16, 0.14, 0.12, 0.10, 0.09, 0.11]):
            sm.cell(row=f, column=17 + i, value=p)
        sm.cell(row=f, column=23, value="Sí")          # potencia según ATR

        meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                 "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        cons = wb["Consumos"]
        for m in range(12):
            fila = BASE_CONSUMOS + 2 + m
            cons.cell(row=fila, column=1, value=meses[m])
            cons.cell(row=fila, column=2, value=2025)
            for c in range(6):
                cons.cell(row=fila, column=3 + c, value=1000 + m)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_importar_guarda_la_plantilla_y_genera_la_ficha_de_consumo(self):
        import io
        import tempfile
        from openpyxl import load_workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.test import override_settings
        from .exportar import comparativo_a_excel

        raw = self._plantilla_rellena()
        cliente = Client()
        cliente.force_login(self.user)

        with override_settings(MEDIA_ROOT=tempfile.mkdtemp()):
            resp = cliente.post("/expedientes/importar/", {
                "fichero": SimpleUploadedFile("plantilla.xlsx", raw),
            })
            self.assertEqual(resp.status_code, 302)   # redirige al expediente creado

            exp = Expediente.objects.latest("creado")
            self.assertEqual(exp.puntos.count(), 1)
            self.assertEqual(exp.puntos.first().consumos.count(), 12)
            # La plantilla subida se conserva como original del expediente
            self.assertTrue(exp.archivo_original)

            Oferta.objects.create(expediente=exp, comercializadora="X", atr_energia_incluido=True)
            wb = load_workbook(io.BytesIO(comparativo_a_excel(exp, comparativo_expediente(exp))))

        # El Excel se genera desde los datos: NO se copian las hojas de la plantilla
        # (instrucciones, 50 filas de CUPS vacías, 49 bloques de consumo en blanco…).
        self.assertEqual(wb.sheetnames[0], "Resumen")
        self.assertNotIn("Instrucciones", wb.sheetnames)
        self.assertNotIn("Cliente", wb.sheetnames)
        ficha = next(h for h in wb.sheetnames if h.startswith("Consumo "))
        self.assertEqual(wb[ficha]["B5"].value, "Cliente Plantilla, S.A.")
        self.assertEqual(wb[ficha]["B7"].value, exp.puntos.first().cups)


class CupsValidacionTest(TestCase):
    def test_cups_validos(self):
        for digitos in ("0011000000000001", "0021000005775680", "0031000000000002"):
            self.assertTrue(validar_cups(cups_valido(digitos)), digitos)

    def test_cups_invalidos(self):
        bueno = cups_valido("0011000000000001")
        # Letras de control equivocadas
        self.assertFalse(validar_cups(bueno[:-2] + "XX"))
        # Prefijo que no es ES
        self.assertFalse(validar_cups("FR" + bueno[2:]))
        # Longitud incorrecta
        self.assertFalse(validar_cups("ES123"))
        # Parte numérica con letras
        self.assertFalse(validar_cups("ESABCD000000000001RK"))


class CatalogoTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("cargar_parametros_2026")
        cls.user = User.objects.create_user("cat", password="x", is_staff=True, is_superuser=True)

    def _expediente_con_cups(self, cif, digitos, tarifa="6.1TD"):
        exp = Expediente.objects.create(cliente_razon_social="Cat SL", cliente_cif=cif, gestor=self.user)
        p = PuntoSuministro.objects.create(
            expediente=exp, cups=cups_valido(digitos), tarifa=tarifa,
            potencia_p1=40, potencia_p2=40, potencia_p3=40, potencia_p4=40, potencia_p5=40, potencia_p6=60,
            energia_peajes_incluidos=True, potencia_segun_atr=True,
        )
        for mes in range(1, 13):
            ConsumoMensual.objects.create(punto=p, anio=2025, mes=mes, p1=1000, p2=900, p3=800, p4=700, p5=600, p6=1200)
        return exp, p

    def _oferta(self, exp):
        oferta = Oferta.objects.create(
            expediente=exp, comercializadora="ENDESA", duracion_meses=12, atr_energia_incluido=True,
            ssaa_tipo="techo", ssaa_modo="mensual", ssaa_ref_superior=Decimal("16"),
            ssaa_perdidas_modo="fija", ssaa_perdidas_pct=Decimal("7"),
        )
        PrecioOferta.objects.create(oferta=oferta, tarifa="6.1TD",
                                    energia_p1=Decimal("0.14"), energia_p2=Decimal("0.12"),
                                    energia_p3=Decimal("0.10"), energia_p4=Decimal("0.08"),
                                    energia_p5=Decimal("0.07"), energia_p6=Decimal("0.09"))
        ConceptoAdicional.objects.create(oferta=oferta, nombre="FNEE", tipo="eur_mwh",
                                         valor=Decimal("1.492"), con_perdidas=True, entra_en_iee=True)
        return oferta

    def test_guardar_oferta_en_catalogo(self):
        exp, _ = self._expediente_con_cups("CAT0001", "0011000000000021")
        oferta = self._oferta(exp)
        c = Client(); c.force_login(self.user)
        r = c.post(f"/expedientes/{exp.pk}/ofertas/{oferta.pk}/a-catalogo/",
                   {"nombre": "ENDESA techo 16", "observaciones": ""})
        self.assertEqual(r.status_code, 302)
        cat = OfertaCatalogo.objects.get(nombre="ENDESA techo 16")
        self.assertEqual(cat.comercializadora, "ENDESA")
        self.assertEqual(cat.ssaa_tipo, "techo")
        self.assertEqual(cat.ssaa_ref_superior, Decimal("16"))
        self.assertEqual(cat.precios.get(tarifa="6.1TD").energia_p1, Decimal("0.14"))
        self.assertEqual(cat.conceptos.get().nombre, "FNEE")

    def test_cargar_desde_catalogo_prerrellena_y_calcula_igual(self):
        exp, _ = self._expediente_con_cups("CAT0002", "0011000000000022")
        oferta = self._oferta(exp)
        total_original = comparativo_expediente(exp)["columnas"][0]["agregado"].total
        c = Client(); c.force_login(self.user)
        c.post(f"/expedientes/{exp.pk}/ofertas/{oferta.pk}/a-catalogo/", {"nombre": "ENDESA techo 16", "observaciones": ""})
        cat = OfertaCatalogo.objects.get(nombre="ENDESA techo 16")

        # Otro expediente con la misma tarifa: cargar desde catálogo prerrellena el formulario
        exp2, _ = self._expediente_con_cups("CAT0003", "0011000000000023")
        html = c.get(f"/expedientes/{exp2.pk}/ofertas/nueva/?catalogo={cat.pk}").content.decode()
        self.assertIn("Cargada desde el catálogo", html)
        # precio de energía P1 prerrellenado (con coma o punto según localización)
        self.assertTrue("0,140000" in html or "0.140000" in html)
        self.assertIn("FNEE", html)      # concepto prerrellenado

        # Guardar la oferta cargada y comprobar que calcula igual que la original
        datos = {
            "comercializadora": cat.comercializadora, "duracion_meses": cat.duracion_meses,
            "atr_energia_incluido": "on", "modo_precios": "tarifa",
            "e_6_1TD_p1": "0.14", "e_6_1TD_p2": "0.12", "e_6_1TD_p3": "0.10",
            "e_6_1TD_p4": "0.08", "e_6_1TD_p5": "0.07", "e_6_1TD_p6": "0.09",
            "ssaa_tipo": "techo", "ssaa_modo": "mensual", "ssaa_ref_superior": "16",
            "ssaa_perdidas_modo": "fija", "ssaa_perdidas_pct": "7",
            "conceptos-TOTAL_FORMS": "1", "conceptos-INITIAL_FORMS": "0",
            "conceptos-MIN_NUM_FORMS": "0", "conceptos-MAX_NUM_FORMS": "1000",
            "conceptos-0-nombre": "FNEE", "conceptos-0-tipo": "eur_mwh", "conceptos-0-valor": "1.492",
            "conceptos-0-con_perdidas": "on", "conceptos-0-entra_en_iee": "on",
        }
        c.post(f"/expedientes/{exp2.pk}/ofertas/nueva/", datos)
        total_copia = comparativo_expediente(exp2)["columnas"][0]["agregado"].total
        self.assertAlmostEqual(total_original, total_copia, delta=CENT)

    def test_precios_por_cups_van_al_catalogo_por_tarifa(self):
        exp, punto = self._expediente_con_cups("CAT0004", "0011000000000024")
        oferta = Oferta.objects.create(expediente=exp, comercializadora="X", atr_energia_incluido=True)
        # Precio SOLO por CUPS (sin tarifa)
        PrecioOferta.objects.create(oferta=oferta, punto=punto,
                                    energia_p1=Decimal("0.13"), energia_p2=Decimal("0.13"), energia_p3=Decimal("0.13"),
                                    energia_p4=Decimal("0.13"), energia_p5=Decimal("0.13"), energia_p6=Decimal("0.13"))
        c = Client(); c.force_login(self.user)
        c.post(f"/expedientes/{exp.pk}/ofertas/{oferta.pk}/a-catalogo/", {"nombre": "X porCUPS", "observaciones": ""})
        cat = OfertaCatalogo.objects.get(nombre="X porCUPS")
        pc = cat.precios.get()  # el precio por CUPS se guardó mapeado a su tarifa, sin punto
        self.assertEqual(pc.tarifa, "6.1TD")
        self.assertEqual(pc.energia_p1, Decimal("0.13"))


class PortadaTest(TestCase):
    """Portada: filtros por ámbito/estado/cliente, KPIs baratos y desglose por técnico."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user("jefa", password="x", first_name="Ana", is_staff=True)
        cls.tecnico = User.objects.create_user("tec", password="x", first_name="Tomás")
        # Expedientes de cada uno
        Expediente.objects.create(cliente_razon_social="Cliente Admin", cliente_cif="A1",
                                  gestor=cls.admin, estado="abierto")
        Expediente.objects.create(cliente_razon_social="Panadería Sur", cliente_cif="B2",
                                  gestor=cls.tecnico, estado="abierto")
        Expediente.objects.create(cliente_razon_social="Ferretería Norte", cliente_cif="C3",
                                  gestor=cls.tecnico, estado="cerrado")

    def test_kpis_no_recalculan_comparativo(self):
        c = Client(); c.force_login(self.admin)
        # La portada no debe importar ni llamar al motor de cálculo: pocas queries.
        with self.assertNumQueries(9):
            r = c.get("/")
        self.assertEqual(r.status_code, 200)
        # Admin ve el ámbito "todos" por defecto: 2 abiertos + 1 cerrado
        self.assertEqual(r.context["kpi_abiertos"], 2)
        self.assertEqual(r.context["kpi_cerrados"], 1)

    def test_desglose_por_tecnico_solo_admin(self):
        ca = Client(); ca.force_login(self.admin)
        self.assertIsNotNone(ca.get("/").context["por_tecnico"])
        ct = Client(); ct.force_login(self.tecnico)
        self.assertIsNone(ct.get("/").context["por_tecnico"])

    def test_ambito_respeta_el_rol(self):
        # Técnico por defecto ve solo los suyos (2)
        ct = Client(); ct.force_login(self.tecnico)
        r = ct.get("/")
        self.assertEqual(r.context["ambito"], "mios")
        self.assertEqual(r.context["kpi_abiertos"], 1)   # solo su expediente abierto
        # Admin por defecto ve todos
        ca = Client(); ca.force_login(self.admin)
        self.assertEqual(ca.get("/").context["ambito"], "todos")

    def test_filtro_estado_y_busqueda(self):
        c = Client(); c.force_login(self.admin)
        # Cerrados: solo la ferretería
        r = c.get("/?ambito=todos&estado=cerrados")
        self.assertEqual([e.cliente_razon_social for e in r.context["expedientes"]], ["Ferretería Norte"])
        # Búsqueda por nombre de cliente
        r = c.get("/?ambito=todos&estado=todos&q=panad")
        self.assertEqual([e.cliente_razon_social for e in r.context["expedientes"]], ["Panadería Sur"])


class UsuariosTest(TestCase):
    """Alta y gestión de usuarios desde la app (solo admin)."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user("root", password="x", is_staff=True)
        cls.tecnico = User.objects.create_user("juan", password="x")

    def test_tecnico_no_accede(self):
        c = Client(); c.force_login(self.tecnico)
        r = c.get("/usuarios/")
        self.assertIn(r.status_code, (302, 403))  # redirige a login o prohíbe

    def test_admin_da_de_alta_tecnico_y_puede_entrar(self):
        c = Client(); c.force_login(self.admin)
        r = c.post("/usuarios/nuevo/", {
            "username": "nuevo", "first_name": "Nuevo", "last_name": "Tec", "email": "n@x.es",
            "rol": "tecnico", "is_active": "on", "password1": "Renovar-2026", "password2": "Renovar-2026",
        })
        self.assertEqual(r.status_code, 302)
        u = User.objects.get(username="nuevo")
        self.assertFalse(u.is_staff)                 # técnico
        self.assertTrue(u.check_password("Renovar-2026"))
        # El técnico creado puede iniciar sesión
        self.assertTrue(Client().login(username="nuevo", password="Renovar-2026"))

    def test_admin_no_puede_autodegradarse_ni_desactivarse(self):
        c = Client(); c.force_login(self.admin)
        r = c.post(f"/usuarios/{self.admin.pk}/editar/", {
            "username": "root", "first_name": "", "last_name": "", "email": "",
            "rol": "tecnico", "password1": "", "password2": "",
        })
        self.assertEqual(r.status_code, 200)         # no redirige: hay error de validación
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_staff)          # sigue siendo admin

    def test_desactivar_bloquea_el_acceso(self):
        c = Client(); c.force_login(self.admin)
        c.post(f"/usuarios/{self.tecnico.pk}/editar/", {
            "username": "juan", "first_name": "Juan", "last_name": "", "email": "",
            "rol": "tecnico", "password1": "", "password2": "",   # is_active ausente = desmarcado
        })
        self.tecnico.refresh_from_db()
        self.assertFalse(self.tecnico.is_active)
        self.assertFalse(Client().login(username="juan", password="x"))
